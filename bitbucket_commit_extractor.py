import requests
import openpyxl
from openpyxl.utils.exceptions import IllegalCharacterError
from datetime import datetime

# Set your Bitbucket Server API URL and project key
API_URL = 'https://ENTER_YOUR_SERVER/rest/api/1.0/projects/BSP'

# Set your Bitbucket Server credentials
USERNAME = ''
PASSWORD = ''

# Set path of output file
EXCEL_FILE = "name.xlsx"

# Set the headers for the Bitbucket Server API requests
headers = {'Content-Type': 'application/json'}

# Authenticate with the Bitbucket Server API
auth = (USERNAME, PASSWORD)

# Send a request to the Bitbucket Server API to get a list of repositories in the project
repo_list = requests.get(API_URL + '/repos', headers=headers, auth=auth).json()

# Get the value of the nextPageStart key in the JSON response
next_page_start = repo_list.get('nextPageStart')

# Keep sending requests to the Bitbucket Server API until all pages of repositories have been retrieved
while next_page_start is not None:
    # Send a request to the Bitbucket Server API to get the next page of repositories in the project
    next_page_repo_list = requests.get(API_URL + f'/repos?start={next_page_start}', headers=headers, auth=auth).json()
    # Add the repositories in the next page to the repository list
    repo_list['values'].extend(next_page_repo_list['values'])
    # Update the value of the nextPageStart key in the JSON response
    next_page_start = next_page_repo_list.get('nextPageStart')

# Create a new Excel workbook and worksheet
workbook = openpyxl.Workbook()
worksheet = workbook.active

# Write the headers for the Excel table
worksheet.cell(row=1, column=1, value='Repository')
worksheet.cell(row=1, column=2, value='Branch')
worksheet.cell(row=1, column=3, value='Commit')
worksheet.cell(row=1, column=4, value='Author')
worksheet.cell(row=1, column=5, value='Date')
worksheet.cell(row=1, column=6, value='Message')

# Initialize the row counter for the Excel table
row = 2

# Iterate over the repositories in the project
for repo in repo_list['values']:
    # Send a request to the Bitbucket Server API to get a list of branches in the repository
    branch_list = requests.get(API_URL + f"/repos/{repo['slug']}/branches", headers=headers, auth=auth).json()
    # Iterate over the branches in the repository
    for branch in branch_list['values']:
        # Send a request to the Bitbucket Server API to get a list of commits in the branch
        commit_list = requests.get(API_URL + f"/repos/{repo['slug']}/commits?until={branch['id']}", headers=headers, auth=auth).json()
        # Iterate over the commits in the branch
        for commit in commit_list['values']:
            # Write the commit information to the Excel table
            worksheet.cell(row=row, column=1, value=repo['name'])
            worksheet.cell(row=row, column=2, value=branch['displayId'])
            worksheet.cell(row=row, column=3, value=commit['id'])
            worksheet.cell(row=row, column=4, value=commit['author']['name'])
            date_time = datetime.fromtimestamp(commit['authorTimestamp']/1000).strftime('%Y-%m-%d %H:%M:%S')
            worksheet.cell(row=row, column=5, value=date_time)
            try:
                worksheet.cell(row=row, column=6, value=commit['message'])
            except IllegalCharacterError:
                worksheet.cell(row=row, column=6, value='')
            # Increment the row counter
            row += 1

# Save the Excel workbook
workbook.save(EXCEL_FILE)